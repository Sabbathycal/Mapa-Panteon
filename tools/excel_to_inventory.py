import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


VALID_STATUSES = {
    "disponible": "disponible",
    "libre": "disponible",

    "separado": "separado",
    "separada": "separado",
    "apartado": "separado",
    "apartada": "separado",

    "vendido": "vendido",
    "vendida": "vendido",

    "utilizado": "utilizado",
    "utilizada": "utilizado",
    "ocupado": "utilizado",
    "ocupada": "utilizado",
    "usado": "utilizado",
    "usada": "utilizado",

    "por construir": "por_construir",
    "por_construir": "por_construir",
    "no construida": "por_construir",
    "no construidas": "por_construir",

    "suspendido": "suspendido",
    "suspendida": "suspendida"
}


REQUIRED_COLUMNS = [
    "tipo",
    "codigo",
    "estatus_venta"
]


def clean(value):
    return str(value or "").strip()


def normalize_header(value):
    return clean(value).lower().replace(" ", "_")


def normalize_status(value):
    raw = clean(value).lower()
    return VALID_STATUSES.get(raw, raw)

def get_final_status(data):
    estatus_ocupacion = clean(data.get("estatus_ocupacion")).lower()
    estatus_venta = clean(data.get("estatus_venta")).lower()

    if estatus_ocupacion in ["ocupado", "ocupada", "utilizado", "utilizada", "usado", "usada"]:
        return "utilizado"

    return normalize_status(estatus_venta)


def normalize_tipo(value):
    return clean(value).lower()


def read_inventory_sheet(excel_path):
    workbook = load_workbook(excel_path, data_only=True)

    if "Inventario_Mapa" not in workbook.sheetnames:
        raise ValueError(
            "No existe la hoja 'Inventario_Mapa' en el Excel. "
            "Crea esa hoja con columnas: tipo, seccion, manzana, zonaId, cara, codigo, estatus."
        )

    sheet = workbook["Inventario_Mapa"]

    headers = [
        normalize_header(cell.value)
        for cell in sheet[1]
    ]

    missing = [
        col
        for col in REQUIRED_COLUMNS
        if col not in headers
    ]

    if missing:
        raise ValueError(
            "Faltan columnas obligatorias en Inventario_Mapa: "
            + ", ".join(missing)
        )

    items = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, row))

        tipo = normalize_tipo(data.get("tipo"))
        codigo = clean(data.get("codigo"))
        estatus = get_final_status(data)

        if not tipo and not codigo and not estatus:
            continue

        if not tipo or not codigo or not estatus:
            print(f"Fila {row_number} ignorada: falta tipo, codigo o estatus.")
            continue

        if tipo not in ["lote", "nicho"]:
            print(f"Fila {row_number} ignorada: tipo no valido ({tipo}).")
            continue

        item = {
            "tipo": tipo,
            "seccion": clean(data.get("seccion")).upper(),
            "manzana": clean(data.get("manzana")).upper(),
            "zonaId": clean(data.get("zonaid") or data.get("zona_id") or data.get("columbario")).upper(),
            "cara": clean(data.get("cara")).lower(),
            "codigo": clean(codigo).upper(),
            "estatus": estatus,
            "referencia_procap": clean(data.get("referencia_procap")),
            "observaciones": clean(data.get("observaciones"))
        }

        items.append(item)

    return items


def main():
    if len(sys.argv) != 3:
        print("Uso:")
        print("python tools/excel_to_inventory.py uploads/PLANOS.xlsx data/inventario-base.json")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo Excel: {excel_path}")

    items = read_inventory_sheet(excel_path)

    payload = {
        "source": "excel",
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
        "items": items
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)

    print(f"Archivo generado: {output_path}")
    print(f"Registros generados: {len(items)}")


if __name__ == "__main__":
    main()